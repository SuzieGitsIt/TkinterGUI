# File:     TkinterGUI_2023-04-07
# Version:  0.0.01
# Author:   Susan Haynes
# Comments/Notes: 
#   (0,0) coordinates are the top left corner of the screen for 1920x1080
#   (0,0) coordinates are the bottom right corner of the screen for 1919x1079
# Online References:

# string   = 'string'     # double quotes
# variable = 'Variable'   # single quotes

from functools import partial
from mimetypes import init                                  # for allowing 015 040 buttons to equal specific values when clicked.
from pathlib import PureWindowsPath                         # library that cleans up windows path extensions
from openpyxl import *                                      # Write to excel
import tkinter as tk                                        # Tkinter's Tk class
import tkinter.ttk as ttk                                   # Tkinter's Tkk class                 
import tkinter.filedialog as tfd
import datetime as dt                                       # Date library
import subprocess                                           # Needed to open an executable
import time                                                 # Needed to call time to count/pause
import psutil,os                                            # Needed for closing an executable
import clr                                                  # Needed to close excel
from PIL import ImageTk, Image                              # Displaying LAL background photo
from tkinter import messagebox                              # Exit standard message box

##########################################################################################################################################
#################################################         1st    GUI SCREEN              #################################################
##########################################################################################################################################
#################################################      INITIALIZING STANDARD DISPLAY     ################################################# 
GUI = tk.Tk()
GUI.title("LAL Measurement")
GUI.geometry('1300x780')                                    # Set the geometry of Tkinter frame
GUI.configure(background = 'white')                             # Set background color
GUI.option_add('*Font', 'Helvetica 12 bold')            # set the font and size for entire GUI
GUI.option_add('*foreground', 'black')                          # set the text color, hex works too '#FFFFFF'
GUI.option_add('*background', 'white')                          # set the background to white

#################################################            BUTTON PRESS STYLE           ################################################ 
style = ttk.Style()
style.theme_use('default')     # alt, default, clam and classic
style.map('T.TButton',background=[('active', 'pressed', 'white'),('!active','white'), ('active','!pressed','grey')]) # active, not active, not pressed
style.map('T.TButton',relief    =[('pressed','sunken'),('!pressed','raised')])  # pressed, not pressed
style.configure('T.TButton', font=('Helvetica', '10'))
style.map('B.TButton',background=[('active', 'pressed', 'white'),('!active','white'), ('active','!pressed','grey')]) # Press me Button always hot pink when pressed
style.map('B.TButton',relief    =[('pressed','sunken'),('!pressed','raised')]) # pressed, not pressed
style.configure('B.TButton', font=('helvetica', '12', 'bold'))
style.map('P.TButton',background=[('active', 'pressed', '#FF69B4'),('!active','white'), ('active','!pressed','grey')]) # Press me Button always hot pink when pressed
style.map('P.TButton',relief    =[('pressed','sunken'),('!pressed','raised')]) # pressed, not pressed
style.configure('P.TButton', font=('helvetica', '12', 'bold'))
style.configure('W.TLabel',  font=('helvetica', '12', 'bold'), foreground= 'black', background= 'white')

# Python is serial, so each widget will output in the order listed below;
#################################################           LAL BACKGROUND IMAGE          ################################################  
def resize_image(event):
    new_width = event.width
    new_height = event.height
    bg_img = copy_img.resize((new_width, new_height))
    new_img = ImageTk.PhotoImage(bg_img)
    lal_img.config(image = new_img)
    lal_img.bg_img = new_img #avoid garbage collection

bg_img = Image.open(r'\\RXS-FS-02\userdocs\shaynes\My Documents\R&D - Software\Python\TkinterGUI_2023-04-06\LAL.png')
copy_img = bg_img.copy()
new_img = ImageTk.PhotoImage(bg_img)
lal_img = ttk.Label(GUI, image = new_img, background='white')
lal_img.bind('<Configure>', resize_image)
lal_img.pack(fill=tk.BOTH, expand = True)

#################################################               EXCEL FILE               ################################################  
wb = Workbook()
ws = wb.active
ws['A1'] = "Credentials"         # Column 1
ws['B1'] = "subWO"               # Column 2
ws['C1'] = "Sample"              # Column 3
ws['D1'] = "Measurement"         # Column 4
ws['E1'] = "Date&Time"           # Column 5
ws['F1'] = "OCT Eq."             # Column 6
ws['G1'] = "Prod/RD"             # Column 7

new_line = ws.max_row + 1

################################################                 MAIN BODY                ################################################   
# Display the command label before the entry box to indicate what information the Opterator is to type
lbl_cmd_date = ttk.Label(GUI, text="Todays Date is:"            , style= 'W.TLabel').place(x=50,y=25)  
lbl_cmd_fold = ttk.Label(GUI, text="Folder Name:"               , style= 'W.TLabel').place(x=50,y=75) 
lbl_cmd_file = ttk.Label(GUI, text="File Name:"                 , style= 'W.TLabel').place(x=50,y=125) 
lbl_cmd_cred = ttk.Label(GUI, text="Enter Operator Credentials:", style= 'W.TLabel').place(x=50,y=175) 
lbl_cmd_WO   = ttk.Label(GUI, text="Enter Work Order Number:"   , style= 'W.TLabel').place(x=50,y=225)   
lbl_cmd_samp = ttk.Label(GUI, text="Enter Sample Sizes:"        , style= 'W.TLabel').place(x=50,y=275)  
lbl_cmd_meas = ttk.Label(GUI, text="Select Measurement Size:"   , style= 'W.TLabel').place(x=50,y=325) 
lbl_cmd_oct  = ttk.Label(GUI, text="Select OCT Equipment:"      , style= 'W.TLabel').place(x=50,y=375)  
lbl_cmd_prd  = ttk.Label(GUI, text="Production:"                , style= 'W.TLabel').place(x=50,y=425)  
lbl_cmd_rnd  = ttk.Label(GUI, text="R&D:"                       , style= 'W.TLabel').place(x=625,y=425)  

# Entry boxes to take information from operator
entry_cred = tk.Entry(GUI , width= 10) 
entry_cred.focus_set()                              # Places cursor in the first entry box.
entry_cred.place(x=300,y=175) 
entry_WO   = tk.Entry(GUI , width= 10) 
entry_WO.place(x=300,y=225) 

# Display the label of what user input as an output
lbl_disp_cred = ttk.Label(GUI, text="Credentials:"       , style= 'W.TLabel').place(x=50, y=520) 
lbl_disp_WO   = ttk.Label(GUI, text="Work Order Number:" , style= 'W.TLabel').place(x=50, y=560) 
lbl_disp_samp = ttk.Label(GUI, text="Sample Sizes:"       , style= 'W.TLabel').place(x=50, y=600) 
lbl_disp_meas = ttk.Label(GUI, text="Measurement Size:"  , style= 'W.TLabel').place(x=50, y=640) 
lbl_disp_oct  = ttk.Label(GUI, text="OCT Equipment:"     , style= 'W.TLabel').place(x=50, y=680) 
lbl_disp_pr   = ttk.Label(GUI, text="Production/R&D:"    , style= 'W.TLabel').place(x=50, y=720) 

# Display the user inputs as outputs 
lbl_out_date = ttk.Label(GUI, text=f'{dt.datetime.now():%b %d, %Y}', style= 'W.TLabel').place(x=300, y=25)
lbl_out_cred = ttk.Label(GUI, text= '', style= 'W.TLabel' , width= 4)
lbl_out_cred.place(x=300, y=520) 
lbl_out_WO   = ttk.Label(GUI, text= '', style= 'W.TLabel' , width= 6)
lbl_out_WO.place(x=300, y=560) 
 
# Display user inputs as outputs
def fun_cred():
    global entry
    cred = entry_cred.get()[:3]                          # entry_cred is the variable we are passing. Limit 3 characters
    lbl_out_cred.configure(text = cred)                  # Display cred entry from user on GUI
    ws.cell(column=1, row=new_line, value = entry_cred.get()[:3])
    print(entry_cred.get()[:3])                          # Print can be removed after developed.

def fun_WO():
    global entry
    WO = entry_WO.get()[:6]                              # entry_WO is the variable we are passing. Limit 10 characters
    lbl_out_WO.configure(text = WO)                      # Display WO entry from user on GUI
    ws.cell(column=2, row=new_line, value = entry_WO.get()[:6])
    print(entry_WO.get()[:6]) 

smpl_sz_1 = 'sample_size_1'
dio_sz_1 = 'dio_size_1'
smpl_sz_2 = 'sample_size_2'
dio_sz_2 = 'dio_size_2'
smpl_sz_3 = 'sample_size_3'
dio_sz_3 = 'dio_size_3'
smpl_sz_4 = 'sample_size_4'
dio_sz_4 = 'dio_size_4'
smpl_sz_5 = 'sample_size_5'
dio_sz_5 = 'dio_size_5'
smpl_sz_6 = 'sample_size_6'
dio_sz_6 = 'dio_size_6'
smpl_sz_7 = 'sample_size_7'
dio_sz_7 = 'dio_size_7'
smpl_sz_8 = 'sample_size_8'
dio_sz_8 = 'dio_size_8'
smpl_sz_9 = 'sample_size_9'
dio_sz_9 = 'dio_size_9'
smpl_sz_10 = 'sample_size_10'
dio_sz_10 = 'dio_size_10'
def fun_samp(entry_samp):
    global smpl_sz_1, dio_sz_1, smpl_sz_2, dio_sz_2, smpl_sz_3, dio_sz_3, smpl_sz_4, dio_sz_4, smpl_sz_5, dio_sz_5
    global smpl_sz_6, dio_sz_6, smpl_sz_7, dio_sz_7, smpl_sz_8, dio_sz_8, smpl_sz_9, dio_sz_9, smpl_sz_10, dio_sz_10
    btn_dio = tk.Entry(GUI, width= 10)
    btn_dio.insert(0,excel_meas+file_meas) 
    btn_dio.place(x=300, y=640)
    #lbl_out_samp.configure(DIO, text = samp)             # Display sample entry from user on Diotpics GUI
    #ws.cell(column=3, row=new_line, value = entry_samp.get()[:3])
 
excel_meas = 'excel_meas_empty'                         # excel_meas is what gets written inside the excel workbook
file_meas  = 'file_meas_empty'                          # fil_meas is what gets used for the excel file name
def fun_meas(entry_meas):
    global excel_meas, file_meas
    if entry_meas== '-B':
        excel_meas = 'Posterior'
        file_meas = '-B'
        btn_015 = tk.Entry(GUI, width= 10)
        btn_015.insert(0,excel_meas+file_meas) 
        btn_015.place(x=300, y=640)
    elif entry_meas== '-A':
        excel_meas = 'Anterior'
        file_meas = '-A'
        btn_040 = tk.Entry(GUI, width= 10)
        btn_040.insert(0,excel_meas+file_meas) 
        btn_040.place(x=300, y=640)
    elif entry_meas== '':
        excel_meas = 'Full Lens'
        file_meas = ''
        btn_100 = tk.Entry(GUI, width= 10)
        btn_100.insert(0,excel_meas+file_meas) 
        btn_100.place(x=300, y=640)
    fun_cred()
    fun_WO()
    print("entry_meas is: ", entry_meas)
    print("excel_meas is: ", excel_meas)
    print("file_meas is: " , file_meas)
    ws.cell(column=4, row=new_line).value = excel_meas

excel_oct  = 'excel_oct_empty'
fold_oct   = 'fold_oct_empty'
eq_num_oct = 'eq_num_oct_empty'
def fun_oct(entry_oct):
    global excel_oct, fold_oct, eq_num_oct
    if entry_oct== 'OCT 1':
        excel_oct  = 'EQ# 1364 '
        fold_oct   = 'OCT 1' 
        eq_num_oct = '1364'
        btn_oct1 = tk.Entry(GUI, width= 15)
        btn_oct1.insert(0,fold_oct+' '+excel_oct) 
        btn_oct1.place(x=300, y=680)
    elif entry_oct== 'OCT 2':
        excel_oct  = 'EQ# 2104 '
        fold_oct   = 'OCT 2' 
        eq_num_oct = '2104'
        btn_oct2 = tk.Entry(GUI, width= 15)
        btn_oct2.insert(0,fold_oct+' '+excel_oct) 
        btn_oct2.place(x=300, y=680)
    ws.cell(column=6, row=new_line).value = excel_oct
    print("entry_oct is: " , entry_oct)
    print("excel_oct is: " , excel_oct)
    print("fold_oct is: "  , fold_oct)
    print("eq_num_oct is: ", eq_num_oct)

excel_pr = 'excel_pr_empty'
file_pr  = 'file_pr_empty'
def fun_prd(entry_pr):
    global excel_pr, file_pr
    if entry_pr== '02':    # Haptics
        excel_pr = 'Haptics'
        file_pr = 'L02-'
        btn_pr02 = tk.Entry(GUI, width= 17)
        btn_pr02.insert(0,file_pr+ ' '+excel_pr) 
        btn_pr02.place(x=300, y=720)
    elif entry_pr== '06':   # R&D
        excel_pr = 'LAL'
        file_pr = 'L06-'
        btn_rd06 = tk.Entry(GUI, width= 17)
        btn_rd06.insert(0,file_pr+ ' '+excel_pr) 
        btn_rd06.place(x=300, y=720)
    elif entry_pr== '07':   # Standard Production
        excel_pr = 'Std Production'
        file_pr = 'L07-'
        btn_pr07 = tk.Entry(GUI, width= 17)
        btn_pr07.insert(0,file_pr+ ' '+excel_pr)   
        btn_pr07.place(x=300, y=720)
    elif entry_pr== '08':   # Next Gen LAL+ R&D
        excel_pr = 'LAL+ R&D'
        file_pr = 'L08-'
        btn_rd08 = tk.Entry(GUI, width= 17)
        btn_rd08.insert(0,file_pr+ ' '+excel_pr) 
        btn_rd08.place(x=300, y=720)
    elif entry_pr== '00':    # Calibration
        excel_pr = 'Calibration'
        file_pr = 'data'
        btn_cal = tk.Entry(GUI, width= 17)
        btn_cal.insert(0,excel_pr) 
        btn_cal.place(x=300, y=720)
    ws.cell(column=7, row=new_line).value = excel_pr
    print("entry_pr is: ", entry_pr)
    print("excel_pr is: ", excel_pr)
    print("file_pr is: " , file_pr)

def fun_save(): 
    ws.cell(column=5, row=new_line).value = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")    
    if excel_pr== 'Calibration':
        filepath = r"\\RXS-FS-02\userdocs\shaynes\My Documents\R&D - Software\Python\VS-TkinterGUI_2023-04-07/" + excel_oct + ' (Lumedica ' + fold_oct + ')' + '/' + dt.datetime.now().strftime("%m-%d-%y") + ' EQ ' + eq_num_oct + ' CALIBRATION'
        filename = 'data_' + dt.datetime.now().strftime("%Y%m%d") + '_' + entry_WO.get()[:6] + '.xlsx'
    else:
        filepath = r"\\RXS-FS-02\userdocs\shaynes\My Documents\R&D - Software\Python\VS-TkinterGUI_2023-04-07/" + excel_oct + '(Lumedica ' + fold_oct + ')'
        filename = file_pr + entry_WO.get()[:6] + file_meas + '.xlsx'
    
    win_filepath = PureWindowsPath(filepath)   
    if not os.path.exists(win_filepath):
        os.makedirs(win_filepath)

    loc = filepath + '/' + filename 
    win_loc = PureWindowsPath(loc)
    wb.save(win_loc)

    lbl_out_fil_pat = ttk.Label(GUI, text= win_filepath, style= 'W.TLabel')
    lbl_out_fil_pat.place(x=300, y=75)
    lbl_out_fil_nam = ttk.Label(GUI, text= filename, style= 'W.TLabel')
    lbl_out_fil_nam.place(x=300, y=125)

    print("Filepath - RAW: \n", filepath)
    print("Filepath - WIN: \n", win_filepath)
    print("Filename:       \n", filename)
    print("Location - RAW: \n", loc)
    print("Location - WIN: \n", win_loc) 

    os.startfile(win_loc)                                                   # Open excel workbook
    time.sleep(20)                                                          # wait 10 seconds to allow to open before force closing
    os.system('TaskKill /F /IM EXCEL.exe')                                  # Force Close ALL excel files

def open_lum():              # Open the calculator, and pause for 5 seconds before executing, this gives the calculator time to open.
    subprocess.Popen('C:\\Windows\\System32\\calc.exe')                     # Open windows calculator
    time.sleep(5)                                                           # wait 5 seconds
    os.system('TaskKill /F /IM CalculatorApp.exe')                          # Force Close windows calculator

def exit_app(): 
    msg_box = tk.messagebox.askquestion('Exit', 'Are you sure you want to exit the application?', icon='warning') 
    if msg_box == 'yes': 
        GUI.destroy() 
    else: 
        tk.messagebox.showinfo('Exit', "Thanks for staying, please continue.") 
 
btn_pres_cnt = 1                                                            # setting count to 0 to be able to call it a global variable within the function
def pink(event):                     
    global btn_pres_cnt                                                  # initializing btn_pres_cnt as a global varaible so that it adds through every iteration
    if(btn_pres_cnt==5 or btn_pres_cnt==10 or btn_pres_cnt==15 or btn_pres_cnt==20 or btn_pres_cnt==25): # button turns pink when btn_pres_cnt=100, and =200 and = 300.
        style.map('T.TButton',background=[('active', 'pressed', '#FF69B4'),('!active','white'), ('active','!pressed','grey')])    # only the button being pressed turns hot pink
        style.configure('T.Button', font= ('Helvetica', '12', 'bold'))
    else:   # else is the normal style
        style.map('T.TButton',background=[('active', 'pressed', 'white'),('!active','white'), ('active','!pressed','grey')])
        style.configure('T.Button', font= ('Helvetica', '12', 'bold'))
    print('btn_pres_cnt = ', btn_pres_cnt)                          
    btn_pres_cnt +=1                                                        # This is always executed at the end of the if else

#################################################        BUTTONS TO BE CLICKED         ################################################   
btn_dio  = ttk.Button(GUI, text='Dioptics', style= 'T.TButton', command=partial(fun_samp, 'Dioptics'))   # Open new GUI for dioptic entry and sample sizes
btn_dio.bind('<Button-1>', pink)
btn_dio.place(x=300,y=272)  

btn_015  = ttk.Button(GUI, text='Posterior', style= 'T.TButton', command=partial(fun_meas, '-B'))       # Post - 015 is the variable we are passing to excel,-B for excel file name
btn_015.bind('<Button-1>', pink)
btn_015.place(x=300,y=322)  

btn_040  = ttk.Button(GUI, text='Anterior',  style= 'T.TButton', command=partial(fun_meas, '-A'))       # Ant - 040 is the variable we are passing to excel, -A for excel file name
btn_040.bind('<Button-1>', pink)
btn_040.place(x=400,y=322) 

btn_100  = ttk.Button(GUI, text='Full Lens', style= 'T.TButton', command=partial(fun_meas, ''))            # Full - 100 is the variable we are passing to excel, blank for the excel file name
btn_100.bind('<Button-1>', pink)
btn_100.place(x=500,y=322) 

btn_oct1  = ttk.Button(GUI, text='OCT 1 EQ# 1364', style= 'T.TButton', command=partial(fun_oct, 'OCT 1'))  # OCT 1 to GUI & filename, EQ# to excel
btn_oct1.bind('<Button-1>', pink)
btn_oct1.place(x=300,y=372) 

btn_oct2  = ttk.Button(GUI, text='OCT 2 EQ# 2104', style= 'T.TButton', command=partial(fun_oct, 'OCT 2'))  # OCT 2 to GUI & filename, EQ# to excel
btn_oct2.bind('<Button-1>', pink)
btn_oct2.place(x=450,y=372) 

btn_pr02  = ttk.Checkbutton(GUI, text= 'Haptics', onvalue= 1, offvalue= 0, style= 'T.TButton', command=partial(fun_prd, '02')) # 02 to GUI & filename, Haptics Production to excel
btn_pr02.bind('<Button-1>', pink)
btn_pr02.place(x=300,y=422) 

btn_pr07  = ttk.Checkbutton(GUI, text= 'Production', onvalue= 1, offvalue= 0, style= 'T.TButton', command=partial(fun_prd, '07')) # 07 to GUI & filename, Standard Production to excel
btn_pr07.bind('<Button-1>', pink)
btn_pr07.place(x=400,y=422) 

btn_rd06  = ttk.Checkbutton(GUI, text= 'LAL', onvalue= 1, offvalue= 0, style= 'T.TButton', command=partial(fun_prd, '06')) # 06 to GUI & filename, R&D LAL to Excel
btn_rd06.bind('<Button-1>', pink)
btn_rd06.place(x=700,y=422) 

btn_rd08  = ttk.Checkbutton(GUI, text= 'LAL+', onvalue= 1, offvalue= 0, style= 'T.TButton', command=partial(fun_prd, '08')) # 08 to GUI & filename, R&D LAL+ to Excel
btn_rd08.bind('<Button-1>', pink)
btn_rd08.place(x=800,y=422) 

btn_cal  = ttk.Checkbutton(GUI, text= 'Calibration', onvalue= 1, offvalue= 0, style= 'T.TButton', command=partial(fun_prd, '00')) # Calibration to Gui and Excel, data to file name
btn_cal.bind('<Button-1>', pink)
btn_cal.place(x=500,y=422) 

btn_sav   = ttk.Button(GUI, text= 'Save',     style= 'B.TButton', command=partial(fun_save))
btn_sav.bind('<Button-1>', pink)
btn_sav.place(x=700, y=720)

btn_lum = ttk.Button(GUI,text='Open Lumedica',style='B.TButton', command=open_lum)              # Currently opens calculator, eventually will open lumedica.exe
btn_lum.bind('<Button-1>', pink)
btn_lum.place(x=810,y=720)

btn_exit = ttk.Button(GUI, text= 'Exit',      style= 'B.TButton', command=exit_app)
btn_exit.bind('<Button-1>', pink)
btn_exit.place(x=950,y=720) 

# Must be at the end of the program in order for the application to run b/c windows is constantly updating
GUI.mainloop()


